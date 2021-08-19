from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from pathlib import Path
import sys

fio_vars = "terse_version_3;fio_version;jobname;groupid;error;read_kb;read_bandwidth;read_iops;read_runtime_ms" \
                   ";read_slat_min;read_slat_max;read_slat_mean;read_slat_dev;read_clat_min;read_clat_max;read_clat_mean" \
                   ";read_clat_dev;read_clat_pct01;read_clat_pct02;read_clat_pct03;read_clat_pct04;read_clat_pct05" \
                   ";read_clat_pct06;read_clat_pct07;read_clat_pct08;read_clat_pct09;read_clat_pct10;read_clat_pct11" \
                   ";read_clat_pct12;read_clat_pct13;read_clat_pct14;read_clat_pct15;read_clat_pct16;read_clat_pct17" \
                   ";read_clat_pct18;read_clat_pct19;read_clat_pct20;read_tlat_min;read_lat_max;read_lat_mean;read_lat_dev" \
                   ";read_bw_min;read_bw_max;read_bw_agg_pct;read_bw_mean;read_bw_dev;write_kb;write_bandwidth;write_iops" \
                   ";write_runtime_ms;write_slat_min;write_slat_max;write_slat_mean;write_slat_dev;write_clat_min" \
                   ";write_clat_max;write_clat_mean;write_clat_dev;write_clat_pct01;write_clat_pct02;write_clat_pct03" \
                   ";write_clat_pct04;write_clat_pct05;write_clat_pct06;write_clat_pct07;write_clat_pct08;write_clat_pct09" \
                   ";write_clat_pct10;write_clat_pct11;write_clat_pct12;write_clat_pct13;write_clat_pct14;write_clat_pct15" \
                   ";write_clat_pct16;write_clat_pct17;write_clat_pct18;write_clat_pct19;write_clat_pct20;write_tlat_min" \
                   ";write_lat_max;write_lat_mean;write_lat_dev;write_bw_min;write_bw_max;write_bw_agg_pct;write_bw_mean" \
                   ";write_bw_dev;cpu_user;cpu_sys;cpu_csw;cpu_mjf;cpu_minf;iodepth_1;iodepth_2;iodepth_4;iodepth_8" \
                   ";iodepth_16;iodepth_32;iodepth_64;lat_2us;lat_4us;lat_10us;lat_20us;lat_50us;lat_100us;lat_250us" \
                   ";lat_500us;lat_750us;lat_1000us;lat_2ms;lat_4ms;lat_10ms;lat_20ms;lat_50ms;lat_100ms;lat_250ms;lat_500ms" \
                   ";lat_750ms;lat_1000ms;lat_2000ms;lat_over_2000ms;disk_name;disk_read_iops;disk_write_iops" \
                   ";disk_read_merges;disk_write_merges;disk_read_ticks;write_ticks;disk_queue_time;disk_util".split(
            ";")

# This is used assuming the fio data folder is in the Linux path of /home/<username>/<fio data folder>
username = "labuser"


class ExcelCreator:
    def __init__(self, excel_name, data_folder_name, is_linux):
        self.data_folder_name = data_folder_name
        self.is_linux = is_linux
        self.windows_path = "C:/Users/rahmaa/PycharmProjects/excelcreator/"
        self.linux_path = "/home/" + username + "/" + self.data_folder_name
        self.excel_name = excel_name
        self.fio_list_of_dicts = []
        self.create_fiodicts()
        self.workbook = Workbook()
        self.create_workbook()

    def create_fiodicts(self):
        # Obtain the path to the data, and then create a list of just files, not directories.
        # Assumes data is in the home directory
        temp_dicts = []
        # For Linux
        if self.is_linux:
            data_folder_path = Path(self.linux_path).glob('**/*')
            data_files = [file for file in data_folder_path if file.is_file()]
            for file in data_files:
                fiodict = self.create_fiodict(file)
                temp_dicts.append(fiodict)
            for dict in temp_dicts:
                if dict:
                    self.fio_list_of_dicts.append(dict)
        # For Windows
        else:
            data_folder_path = Path(self.windows_path).glob('**/*.txt')
            data_files = [str(file) for file in data_folder_path]
            for file in data_files:
                fiodict = self.create_fiodict(file)
                temp_dicts.append(fiodict)
            for dict in temp_dicts:
                if dict:
                    self.fio_list_of_dicts.append(dict)

    def create_fiodict(self, file):
        fio_dict = {}

        # Only open relevant FIO files, which will have the bs string in the file name
        # Create dictionary of FIO vars and values

        if "bs" in str(file):
            with open(str(file), 'r') as reader:
                drive_data = str(file).split("_")
                model = drive_data[0].split("/")[-1]
                serial = drive_data[1]
                fw_rev = drive_data[2]
                timestamp = (drive_data[3] + " " + drive_data[4]).split("/")[0]
                fio_data = reader.readline().split(";")
                fio_data[-1] = fio_data[-1].rstrip("\n")
                job_info = fio_data[2].split("_")
                test_type = job_info[0]
                ioengine = job_info[1]
                threads = job_info[2][1:]
                iodepth = job_info[3][2:]
                bs = job_info[4][2:]
                bs = bs.replace("k", "")
                cpu_model = drive_data[5]
                server_model = drive_data[6].split("/")[0]

                fio_dict["model"] = model
                fio_dict["serial"] = serial
                fio_dict["fw_rev"] = fw_rev
                fio_dict["timestamp"] = timestamp
                fio_dict["test_type"] = test_type
                fio_dict["threads"] = threads
                fio_dict["bs"] = bs
                fio_dict["io_engine"] = ioengine
                fio_dict["iodepth"] = iodepth
                fio_dict["cpu_model"] = cpu_model
                fio_dict["server_model"] = server_model

                for i in range(len(fio_vars)):
                    fio_dict[fio_vars[i]] = fio_data[i]
        # for k, v in fio_dict.items():
        # print(k, v)
        return fio_dict

    def create_workbook(self):
        # Create two sheets, one for clean data and one for raw output
        ws = self.workbook.active
        ws.title = "FIO Drive Data"
        raw_output = self.workbook.create_sheet("Raw Output")

        # Create labels for sheets
        self.input_raw_output_labels(raw_output)
        self.input_drive_data_labels(ws)

        # Input data into sheets
        self.input_raw_output_data(raw_output)
        self.input_drive_data(ws, raw_output)

        # Save workbook
        self.workbook.save(self.excel_name)

    def input_raw_output_labels(self, worksheet):
        i = 0
        for col in worksheet.iter_cols(min_row=1, max_col=130, max_row=1):
            for cell in col:
                cell.value = fio_vars[i]
                i += 1

    def input_drive_data_labels(self, worksheet):
        drive_info = self.data_folder_name.split("_")
        
        worksheet.column_dimensions['A'].width = 49.55
        worksheet.column_dimensions['B'].width = 39.36
        worksheet.column_dimensions['C'].width = 46.91

        worksheet['A1'] = "Specifications"
        worksheet['C1'] = drive_info[0].replace("-", " ")
        worksheet['C3'] = drive_info[0]
        worksheet['C4'] = drive_info[1]
        worksheet['C11'] = drive_info[6].replace("/", "").replace("-", " ")
        worksheet['C12'] = drive_info[5].replace("-", " ")
        worksheet['A2'] = "Capacity"
        worksheet['A3'] = "Standard Model Number"
        worksheet['A4'] = "Standard Serial Number"
        worksheet['A5'] = "Features"
        worksheet['A6'] = "Interface"
        worksheet['A7'] = "NAND Flash Type"
        worksheet['A8'] = "Form Factor"
        worksheet['A9'] = "DWPD"
        worksheet['A10'] = "Power (W)"
        worksheet['A11'] = "Platform used for Testing"
        worksheet['A12'] = "CPU"

        worksheet.merge_cells('A14:A16')
        worksheet.merge_cells('A17:A19')
        worksheet.merge_cells('A20:A22')
        worksheet.merge_cells('A23:A25')
        worksheet.merge_cells('A26:A32')
        worksheet.merge_cells('A33:A39')
        worksheet.merge_cells('A40:A46')

        worksheet.merge_cells('A48:A50')
        worksheet.merge_cells('A51:A53')
        worksheet.merge_cells('A54:A56')
        worksheet.merge_cells('A57:A59')
        worksheet.merge_cells('A60:A66')
        worksheet.merge_cells('A67:A73')
        worksheet.merge_cells('A74:A80')

        worksheet['A14'] = "Sequential Read (MB/s) Sustained, 128KB, QD256, T1"
        worksheet['A17'] = "Sequential Write (MB/s) Sustained, 128KB, QD256, T1"
        worksheet['A20'] = "Random Read (IOPS) Sustained, 4KB, QD256, T1"
        worksheet['A23'] = "Random Write (IOPS) Sustained, 4KB, QD256, T1"
        worksheet['A26'] = "Random 30% Write (IOPS) Sustained, 4KB, QD256, T1"
        worksheet['A33'] = "Random 50% Write (IOPS) Sustained, 4KB, QD256, T1"
        worksheet['A40'] = "Random 70% Write (IOPS) Sustained, 4KB, QD256, T1"

        worksheet['A48'] = "Sequential Read (MB/s) Sustained, 128KB, QD256, T8"
        worksheet['A51'] = "Sequential Write (MB/s) Sustained, 128KB, QD256, T8"
        worksheet['A54'] = "Random Read (IOPS) Sustained, 4KB, QD256, T8"
        worksheet['A57'] = "Random Write (IOPS) Sustained, 4KB, QD256, T8"
        worksheet['A60'] = "Random 30% Write (IOPS) Sustained, 4KB, QD256, T8"
        worksheet['A67'] = "Random 50% Write (IOPS) Sustained, 4KB, QD256, T8"
        worksheet['A74'] = "Random 70% Write (IOPS) Sustained, 4KB, QD256, T8"

        align_cells = ['A14', 'A17', 'A20', 'A23', 'A26', 'A33', 'A40', 'A48', 'A51', 'A54', 'A57', 'A60', 'A67', 'A74',
                       'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9', 'C10', 'C11', 'C12', 'C13', 'C47']
        for cell in align_cells:
            worksheet[cell].alignment = Alignment(horizontal='center', vertical='center')

        worksheet['A13'] = "Performance for 1 Thread"
        worksheet['B13'] = "Test Metrics"
        worksheet['B14'] = "Throughput (MB/s)"
        worksheet['B15'] = "IOPS"
        worksheet['B16'] = "Average Read Latency (usec)"
        worksheet['B17'] = "Throughput (MB/s)"
        worksheet['B18'] = "IOPS"
        worksheet['B19'] = "Average Write Latency (usec)"
        worksheet['B20'] = "Throughput (MB/s)"
        worksheet['B21'] = "IOPS"
        worksheet['B22'] = "Average Read Latency (usec)"
        worksheet['B23'] = "Throughput (MB/s)"
        worksheet['B24'] = "IOPS"
        worksheet['B25'] = "Average Write Latency (usec)"
        worksheet['B26'] = "Read Throughput (MB/s)"
        worksheet['B27'] = "Write Throughput (MB/s)"
        worksheet['B28'] = "Read IOPS"
        worksheet['B29'] = "Write IOPS"
        worksheet['B30'] = "Total IOPS"
        worksheet['B31'] = "Average Read Latency (usec)"
        worksheet['B32'] = "Average Write Latency (usec)"
        worksheet['B33'] = "Read Throughput (MB/s)"
        worksheet['B34'] = "Write Throughput (MB/s)"
        worksheet['B35'] = "Read IOPS"
        worksheet['B36'] = "Write IOPS"
        worksheet['B37'] = "Total IOPS"
        worksheet['B38'] = "Average Read Latency (usec)"
        worksheet['B39'] = "Average Write Latency (usec)"
        worksheet['B40'] = "Read Throughput (MB/s)"
        worksheet['B41'] = "Write Throughput (MB/s)"
        worksheet['B42'] = "Read IOPS"
        worksheet['B43'] = "Write IOPS"
        worksheet['B44'] = "Total IOPS"
        worksheet['B45'] = "Average Read Latency (usec)"
        worksheet['B46'] = "Average Write Latency (usec)"
        worksheet['C13'] = "1 FIO Thread"
        worksheet['A47'] = "Performance for 8 Threads"
        worksheet['B47'] = "Test Metrics"
        worksheet['B48'] = "Throughput (MB/s)"
        worksheet['B49'] = "IOPS"
        worksheet['B50'] = "Average Read Latency (usec)"
        worksheet['B51'] = "Throughput (MB/s)"
        worksheet['B52'] = "IOPS"
        worksheet['B53'] = "Average Write Latency (usec)"
        worksheet['B54'] = "Throughput (MB/s)"
        worksheet['B55'] = "IOPS"
        worksheet['B56'] = "Average Read Latency (usec)"
        worksheet['B57'] = "Throughput (MB/s)"
        worksheet['B58'] = "IOPS"
        worksheet['B59'] = "Average Write Latency (usec)"
        worksheet['B60'] = "Read Throughput (MB/s)"
        worksheet['B61'] = "Write Throughput (MB/s)"
        worksheet['B62'] = "Read IOPS"
        worksheet['B63'] = "Write IOPS"
        worksheet['B64'] = "Total IOPS"
        worksheet['B65'] = "Average Read Latency (usec)"
        worksheet['B66'] = "Average Write Latency (usec)"
        worksheet['B67'] = "Read Throughput (MB/s)"
        worksheet['B68'] = "Write Throughput (MB/s)"
        worksheet['B69'] = "Read IOPS"
        worksheet['B70'] = "Write IOPS"
        worksheet['B71'] = "Total IOPS"
        worksheet['B72'] = "Average Read Latency (usec)"
        worksheet['B73'] = "Average Write Latency (usec)"
        worksheet['B74'] = "Read Throughput (MB/s)"
        worksheet['B75'] = "Write Throughput (MB/s)"
        worksheet['B76'] = "Read IOPS"
        worksheet['B77'] = "Write IOPS"
        worksheet['B78'] = "Total IOPS"
        worksheet['B79'] = "Average Read Latency (usec)"
        worksheet['B80'] = "Average Write Latency (usec)"
        worksheet['C47'] = "8 FIO Threads"

        self.set_fill(worksheet, 'A1:C1', "black")
        self.set_font(worksheet, 'A1:C1', "white")
        self.set_fill(worksheet, 'A5:C5', "gray")
        self.set_font(worksheet, 'A5:C5', "white")
        self.set_fill(worksheet, 'A11:C11', "gray")
        self.set_font(worksheet, 'A11:C11', "white")
        self.set_fill(worksheet, 'A13:C13', "gray")
        self.set_font(worksheet, 'A13:C13', "white")
        self.set_fill(worksheet, 'A47:C47', "gray")
        self.set_font(worksheet, 'A47:C47', "white")

        self.set_fill(worksheet, 'A17:A19', "lightgray")
        self.set_fill(worksheet, 'B17:B19', "lightgray")
        self.set_fill(worksheet, 'C17:C19', "lightgray")
        self.set_fill(worksheet, 'A23:A25', "lightgray")
        self.set_fill(worksheet, 'B23:B25', "lightgray")
        self.set_fill(worksheet, 'C23:C25', "lightgray")
        self.set_fill(worksheet, 'A33:A39', "lightgray")
        self.set_fill(worksheet, 'B33:B39', "lightgray")
        self.set_fill(worksheet, 'C33:C39', "lightgray")
        self.set_fill(worksheet, 'A51:A53', "lightgray")
        self.set_fill(worksheet, 'B51:B53', "lightgray")
        self.set_fill(worksheet, 'C51:C53', "lightgray")
        self.set_fill(worksheet, 'A57:A59', "lightgray")
        self.set_fill(worksheet, 'B57:B59', "lightgray")
        self.set_fill(worksheet, 'C57:C59', "lightgray")
        self.set_fill(worksheet, 'A67:A73', "lightgray")
        self.set_fill(worksheet, 'B67:B73', "lightgray")
        self.set_fill(worksheet, 'C67:C73', "lightgray")

        self.set_border(worksheet, 'C1:C80', "left")
        self.set_border(worksheet, 'D1:D80', "left")
        self.set_border(worksheet, 'A5:C5', "topbot")
        self.set_border(worksheet, 'A11:C11', "topbot")
        self.set_border(worksheet, 'A13:C13', "topbot")
        self.set_border(worksheet, 'A16:C16', "bot")
        self.set_border(worksheet, 'A19:C19', "bot")
        self.set_border(worksheet, 'A22:C22', "bot")
        self.set_border(worksheet, 'A25:C25', "bot")
        self.set_border(worksheet, 'A32:C32', "bot")
        self.set_border(worksheet, 'A39:C39', "bot")
        self.set_border(worksheet, 'A46:C46', "bot")
        self.set_border(worksheet, 'A47:C47', "bot")
        self.set_border(worksheet, 'A50:C50', "bot")
        self.set_border(worksheet, 'A53:C53', "bot")
        self.set_border(worksheet, 'A56:C56', "bot")
        self.set_border(worksheet, 'A59:C59', "bot")
        self.set_border(worksheet, 'A66:C66', "bot")
        self.set_border(worksheet, 'A73:C73', "bot")
        self.set_border(worksheet, 'A80:C80', "bot")
        self.set_border(worksheet, 'C16:C16', "botleft")
        self.set_border(worksheet, 'C19:C19', "botleft")
        self.set_border(worksheet, 'C22:C22', "botleft")
        self.set_border(worksheet, 'C25:C25', "botleft")
        self.set_border(worksheet, 'C32:C32', "botleft")
        self.set_border(worksheet, 'C39:C39', "botleft")
        self.set_border(worksheet, 'C46:C46', "botleft")
        self.set_border(worksheet, 'C50:C50', "botleft")
        self.set_border(worksheet, 'C53:C53', "botleft")
        self.set_border(worksheet, 'C56:C56', "botleft")
        self.set_border(worksheet, 'C59:C59', "botleft")
        self.set_border(worksheet, 'C66:C66', "botleft")
        self.set_border(worksheet, 'C73:C73', "botleft")
        self.set_border(worksheet, 'C80:C80', "botleft")

    def input_raw_output_data(self, worksheet):
        i = 0
        test_row = 2
        for fio_dict in self.fio_list_of_dicts:
            for col in worksheet.iter_cols(min_row=test_row, max_col=130, max_row=test_row):
                for cell in col:
                    cell.value = fio_dict[fio_vars[i]]
                    i += 1
            test_row += 1
            i = 0

    def input_drive_data(self, clean_data_ws, raw_data_ws):
        i = 2
        while i < 16:
            job_info = str(raw_data_ws['C' + str(i)].value).split("_")
            job_type = job_info[0]
            threads = job_info[2]
            qd = job_info[3].upper()
            if "seqwrite" == job_type and "t1" == threads:
                clean_data_ws['A17'] = "Sequential Write (MB/s) Sustained, 128KB, %s, T1" % qd
                clean_data_ws['C17'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C18'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C19'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "seqwrite" == job_type and "t8" == threads:
                clean_data_ws['A51'] = "Sequential Write (MB/s) Sustained, 128KB, %s, T8" % qd
                clean_data_ws['C51'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C52'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C53'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "seqread" == job_type and "t1" == threads:
                clean_data_ws['A14'] = "Sequential Read (MB/s) Sustained, 128KB, %s, T1" % qd
                clean_data_ws['C14'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C15'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C16'] = round(float(raw_data_ws['P' + str(i)].value), 0)
            elif "seqread" == job_type and "t8" == threads:
                clean_data_ws['A48'] = "Sequential Read (MB/s) Sustained, 128KB, %s, T8" % qd
                clean_data_ws['C48'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C49'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C50'] = round(float(raw_data_ws['P' + str(i)].value), 0)
            elif "randwrite" == job_type and "t1" == threads:
                clean_data_ws['A23'] = "Random Write (IOPS) Sustained, 4KB, %s, T1" % qd
                clean_data_ws['C23'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C24'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C25'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "randwrite" == job_type and "t8" == threads:
                clean_data_ws['A57'] = "Random Write (IOPS) Sustained, 4KB, %s, T8" % qd
                clean_data_ws['C57'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C58'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C59'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "randread" == job_type and "t1" == threads:
                clean_data_ws['A20'] = "Random Read (IOPS) Sustained, 4KB, %s, T1" % qd
                clean_data_ws['C20'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C21'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C22'] = round(float(raw_data_ws['P' + str(i)].value), 0)
            elif "randread" == job_type and "t8" == threads:
                clean_data_ws['A54'] = "Random Read (IOPS) Sustained, 4KB, %s, T8" % qd
                clean_data_ws['C54'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C55'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C56'] = round(float(raw_data_ws['P' + str(i)].value), 0)
            elif "randmixedread70write30" == job_type and "t1" == threads:
                clean_data_ws['A26'] = "Random 30%% Write (IOPS) Sustained, 4KB, %s, T1" % qd
                clean_data_ws['C26'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C27'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C28'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C29'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C30'] = clean_data_ws['C28'].value + clean_data_ws['C29'].value
                clean_data_ws['C31'] = round(float(raw_data_ws['P' + str(i)].value), 0)
                clean_data_ws['C32'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "randmixedread70write30" == job_type and "t8" == threads:
                clean_data_ws['A60'] = "Random 30%% Write (IOPS) Sustained, 4KB, %s, T8" % qd
                clean_data_ws['C60'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C61'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C62'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C63'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C64'] = clean_data_ws['C62'].value + clean_data_ws['C63'].value
                clean_data_ws['C65'] = round(float(raw_data_ws['P' + str(i)].value), 0)
                clean_data_ws['C66'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "randmixedread50write50" == job_type and "t1" == threads:
                clean_data_ws['A33'] = "Random 50%% Write (IOPS) Sustained, 4KB, %s, T1" % qd
                clean_data_ws['C33'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C34'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C35'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C36'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C37'] = clean_data_ws['C35'].value + clean_data_ws['C36'].value
                clean_data_ws['C38'] = round(float(raw_data_ws['P' + str(i)].value), 0)
                clean_data_ws['C39'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "randmixedread50write50" == job_type and "t8" == threads:
                clean_data_ws['A67'] = "Random 50%% Write (IOPS) Sustained, 4KB, %s, T8" % qd
                clean_data_ws['C67'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C68'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C69'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C70'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C71'] = clean_data_ws['C69'].value + clean_data_ws['C70'].value
                clean_data_ws['C72'] = round(float(raw_data_ws['P' + str(i)].value), 0)
                clean_data_ws['C73'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "randmixedread30write70" == job_type and "t1" == threads:
                clean_data_ws['A40'] = "Random 70%% Write (IOPS) Sustained, 4KB, %s, T1" % qd
                clean_data_ws['C40'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C41'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C42'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C43'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C44'] = clean_data_ws['C42'].value + clean_data_ws['C43'].value
                clean_data_ws['C45'] = round(float(raw_data_ws['P' + str(i)].value), 0)
                clean_data_ws['C46'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            elif "randmixedread30write70" == job_type and "t8" == threads:
                clean_data_ws['A74'] = "Random 70%% Write (IOPS) Sustained, 4KB, %s, T8" % qd
                clean_data_ws['C74'] = round(int(raw_data_ws['G' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C75'] = round(int(raw_data_ws['AV' + str(i)].value) * 0.001024, 0)
                clean_data_ws['C76'] = int(raw_data_ws['H' + str(i)].value)
                clean_data_ws['C77'] = int(raw_data_ws['AW' + str(i)].value)
                clean_data_ws['C78'] = clean_data_ws['C76'].value + clean_data_ws['C77'].value
                clean_data_ws['C79'] = round(float(raw_data_ws['P' + str(i)].value), 0)
                clean_data_ws['C80'] = round(float(raw_data_ws['BE' + str(i)].value), 0)
            i += 1

    def set_border(self, ws, cell_range, border_type):
        if border_type == "topbot":
            border = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        elif border_type == "left":
            border = Border(left=Side(border_style='thin'))
        elif border_type == "topbotleft":
            border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'))
        elif border_type == "botleft":
            border = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'))
        elif border_type == "bot":
            border = Border(bottom=Side(border_style='thin'))
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border

    def set_fill(self, ws, cell_range, color):
        if color == "black":
            fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
        elif color == "gray":
            fill = PatternFill(start_color='757171', end_color='757171', fill_type='solid')
        elif color == "lightgray":
            fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        else:
            fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        cols = ws[cell_range]
        for col in cols:
            for cell in col:
                cell.fill = fill

    def set_font(self, ws, cell_range, color):
        if color == "white":
            font = Font(color="00FFFFFF")
        else:
            font = Font(color="00000000")
        cols = ws[cell_range]
        for col in cols:
            for cell in col:
                cell.font = font


if __name__ == "__main__":
    #data_folder_name = str(sys.argv[1])
    data_folder_name = "Dell-Ent-NVMe-CM6-RI_11K0A00RTCE7_2.1.5_20210818_155702_AMD-EPYC-7552-48-Core-Processor_PowerEdge-R7525"
    time_stamp = str(data_folder_name).split("_")[3] + "_" + str(data_folder_name).split("_")[4]
    excel_name = str(data_folder_name.split("_")[0]) + "_FIO_FourCorners_Data_" + time_stamp + ".xlsx"
    ExcelCreator(excel_name, data_folder_name, False)
